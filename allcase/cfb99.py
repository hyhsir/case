# __Author:HYH
# DATE:2019/3/28

# 九九乘法表
"""
2种方法
重点理解如下：
for j in range（1，i+1）
当i = 1，j=(1,2)，此时j 的取值就是1
当i = 2，j=(1,3)，此时j 的取值就是1，2
当i = 3，j=(1,4)，此时j 的取值就是1,2,3
当i = 4，j=(1,5)，此时j 的取值就是1,2,3,4
当i = 5，j=(1,6)，此时j 的取值就是1,2,3,4,5
...

end=' '意思是末尾不换行，加空格。print默认是打印一行，结尾有换行。

str.format()格式化字符串的函数，基本语法是通过 {} 和 : 来代替以前的 %
format 函数可以接受不限个参数，位置可以不按顺序
"""
# for i in range(1, 10):
#     for j in range(1, i + 1):
#         print('{}x{}={}\t'.format(j, i, i*j), end='')
#         # print('{0}x{1}={2}\t'.format(j, i, i*j), end='')
#         # print("%dx%d=%d" % (j, i, i * j), end=' ')
#     print()


from openpyxl import Workbook

# 在内存中创建一个workbook对象，而且会至少创建一个worksheet
wb = Workbook()

ws = wb.get_active_sheet()
print(ws.title)
ws.title = '九九乘法表'  # 设置worksheet的标题
print(ws.title)

# ws.cell(row=3, column=1, value=6)  # 设置单元格的值
# 把九九乘法表写入单元格
for r in range(1, 10):
    for c in range(1, r+1):
        ws.cell(row=r, column=c, value="%dx%d=%d" % (c, r, r * c))

# 创建新sheet，写入数据
new_ws = wb.create_sheet(title='new_sheet')
for row in range(1, 101):
    for col in range(1, 11):
        # 不推荐这样写入数据
        new_ws.cell(row=row, column=col).value = row + col

# 最后一定要保存！
wb.save(filename='write.xlsx')



















